from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib import messages
from django.db.models import Sum, Q
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
import json
import uuid
from .models import ContaPagar, ContaReceber, Unidade, Categoria, PerfilUsuario, ContaBancaria
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def parse_currency(val_str):
    if not val_str:
        return Decimal("0.00")
    val_str = str(val_str).replace("R$", "").replace(" ", "").strip()
    val_str = val_str.replace(".", "").replace(",", ".")
    try:
        return Decimal(val_str)
    except Exception:
        return Decimal("0.00")

def calcular_meses_datas(data_inicio, data_fim):
    datas = []
    ano = data_inicio.year
    mes = data_inicio.month
    dia_original = data_inicio.day

    while True:
        dia_ajustado = min(dia_original, 28)
        dt = datetime(ano, mes, dia_ajustado).date()
        if dt > data_fim:
            break
        datas.append(dt)
        if mes == 12:
            mes = 1
            ano += 1
        else:
            mes += 1
    return datas

def gerar_proximas_datas(data_inicio, qtd_meses=24):
    datas = []
    ano = data_inicio.year
    mes = data_inicio.month
    dia_original = data_inicio.day

    for _ in range(qtd_meses):
        dia_ajustado = min(dia_original, 28)
        dt = datetime(ano, mes, dia_ajustado).date()
        datas.append(dt)
        if mes == 12:
            mes = 1
            ano += 1
        else:
            mes += 1
    return datas

def login_view(request):
    if request.user.is_authenticated:
        return redirect("dashboard_gestor")
    if request.method == "POST":
        u = request.POST.get("username")
        p = request.POST.get("password")
        user = authenticate(request, username=u, password=p)
        if user is not None:
            login(request, user)
            return redirect("dashboard_gestor")
        else:
            messages.error(request, "Usuário ou senha incorretos.")
    return render(request, "financeiro/login.html")

def logout_view(request):
    logout(request)
    return redirect("login")

@login_required
def alterar_senha_view(request):
    if request.method == "POST":
        senha_atual = request.POST.get("senha_atual")
        nova_senha = request.POST.get("nova_senha")
        confirma_senha = request.POST.get("confirma_senha")

        user = request.user
        if not user.check_password(senha_atual):
            messages.error(request, "A senha atual está incorreta.")
        elif not nova_senha or nova_senha != confirma_senha:
            messages.error(request, "A nova senha e a confirmação não conferem.")
        else:
            user.set_password(nova_senha)
            user.save()
            update_session_auth_hash(request, user)
            messages.success(request, "Senha alterada com sucesso!")
    return redirect(request.META.get("HTTP_REFERER", "dashboard_gestor"))

@login_required
def dashboard_gestor(request):
    user = request.user
    perfil = getattr(user, "perfil", None)
    if not perfil:
        try:
            perfil, _ = PerfilUsuario.objects.get_or_create(
                user=user,
                defaults={
                    "papel": "GESTOR" if (user.is_superuser or user.is_staff) else "OPERADOR",
                    "polos_ids": ""
                }
            )
        except Exception:
            class DummyPerfil:
                papel = "GESTOR" if (user.is_superuser or user.is_staff) else "OPERADOR"
                polos_ids = ""
                def get_unidades_ids(self): return []
            perfil = DummyPerfil()

    is_operador = (perfil.papel == "OPERADOR")
    is_controller = (perfil.papel == "CONTROLLER")
    is_gestor = (perfil.papel == "GESTOR" or user.is_superuser or user.is_staff)
    pode_configurar = (is_gestor or is_controller)

    ids_permitidos = perfil.get_unidades_ids()
    unidades_permitidas = Unidade.objects.filter(id__in=ids_permitidos, ativo=True) if ids_permitidos else Unidade.objects.none()
    has_polos_vinculados = bool(ids_permitidos)

    unidades_get = request.GET.getlist("unidades")
    unidade_unica = request.GET.get("unidade")
    if unidade_unica and not unidades_get:
        unidades_get = [unidade_unica]

    periodo = request.GET.get("periodo", "todos")
    data_de = request.GET.get("data_de")
    data_ate = request.GET.get("data_ate")
    categoria_id = request.GET.get("categoria_filtro")
    conta_origem_filtro = request.GET.get("conta_origem_filtro")
    
    ano_param = request.GET.get("ano")
    mes_param = request.GET.get("mes")

    hoje = timezone.localdate()
    if ano_param and ano_param.isdigit() and mes_param and mes_param.isdigit():
        ano_atual = int(ano_param)
        mes_atual = int(mes_param)
    else:
        ano_atual = hoje.year
        mes_atual = hoje.month

    meses_nomes = {
        1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
    }
    nome_mes_atual = f"{meses_nomes.get(mes_atual, 'Mês')} {ano_atual}"

    if mes_atual == 1:
        mes_ant, ano_ant = 12, ano_atual - 1
    else:
        mes_ant, ano_ant = mes_atual - 1, ano_atual

    if mes_atual == 12:
        mes_prox, ano_prox = 1, ano_atual + 1
    else:
        mes_prox, ano_prox = mes_atual + 1, ano_atual

    nomes_uninter = ["Piracicaba", "Americana", "Sumaré", "Sumaré 2", "Rio Claro", "Araras", "Leme", "Conchal", "Pirassununga", "Matão", "Araraquara", "Olímpia", "Luciana Pessoal", "Rangélia Pessoal", "Alexandre", "Eletroposto Rosario"]

    for nome in nomes_uninter:
        Unidade.objects.get_or_create(nome=nome, defaults={"ativo": True})

    todas_unidades = list(Unidade.objects.filter(ativo=True))
    unidades_dict = {u.nome.lower(): u for u in todas_unidades}

    def map_unidades(nomes):
        res = []
        for n in nomes:
            u = unidades_dict.get(n.lower())
            if u:
                if is_gestor or not has_polos_vinculados or u.id in ids_permitidos:
                    res.append(u)
        return res

    todas_unidades_mapeadas = map_unidades(nomes_uninter)

    despesas = ContaPagar.objects.select_related("unidade", "categoria", "conta_origem", "criado_por").all()
    receitas = ContaReceber.objects.select_related("unidade", "categoria", "conta_destino").all()

    if is_operador:
        receitas = ContaReceber.objects.none()
        if has_polos_vinculados:
            despesas = despesas.filter(unidade_id__in=ids_permitidos, criado_por=user)
            todas_unidades = list(unidades_permitidas)
        else:
            despesas = despesas.none()
    elif is_controller:
        if has_polos_vinculados:
            despesas = despesas.filter(unidade_id__in=ids_permitidos)
            receitas = receitas.filter(unidade_id__in=ids_permitidos)
            todas_unidades = list(unidades_permitidas)
        else:
            despesas = despesas.none()
            receitas = receitas.none()

    if unidades_get:
        ids_filtrar = [int(uid) for uid in unidades_get if str(uid).isdigit()]
        if ids_filtrar:
            despesas = despesas.filter(unidade_id__in=ids_filtrar)
            receitas = receitas.filter(unidade_id__in=ids_filtrar)

    despesas = despesas.filter(data_vencimento__year=ano_atual, data_vencimento__month=mes_atual)
    receitas = receitas.filter(data_previsao__year=ano_atual, data_previsao__month=mes_atual)

    if conta_origem_filtro and conta_origem_filtro.isdigit():
        despesas = despesas.filter(conta_origem_id=int(conta_origem_filtro))

    if periodo == "hoje":
        despesas = despesas.filter(data_vencimento=hoje)
        receitas = receitas.filter(data_previsao=hoje)
    elif periodo == "semana":
        fim_semana = hoje + timedelta(days=7)
        despesas = despesas.filter(data_vencimento__range=[hoje, fim_semana])
        receitas = receitas.filter(data_previsao__range=[hoje, fim_semana])
    elif periodo == "atrasados":
        despesas = despesas.filter(data_vencimento__lt=hoje).exclude(status="PAGO")
    elif periodo == "pagos":
        despesas = despesas.filter(status="PAGO")
    elif periodo == "aprovados":
        despesas = despesas.filter(status="APROVADO")
    elif data_de or data_ate:
        if data_de:
            despesas = despesas.filter(data_vencimento__gte=data_de)
            receitas = receitas.filter(data_previsao__gte=data_de)
        if data_ate:
            despesas = despesas.filter(data_vencimento__lte=data_ate)
            receitas = receitas.filter(data_previsao__lte=data_ate)

    if categoria_id and categoria_id.isdigit():
        despesas = despesas.filter(categoria_id=int(categoria_id))
        receitas = receitas.filter(categoria_id=int(categoria_id))

    total_despesas = despesas.aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")
    total_custo_fixo = despesas.filter(recorrente=True).aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")
    total_pendente = despesas.filter(status="PENDENTE").aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")
    total_aprovado = despesas.filter(status="APROVADO").aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")
    total_pago = despesas.filter(status="PAGO").aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")

    total_receitas = receitas.aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")
    saldo_projetado = total_receitas - total_despesas

    try:
        if is_gestor:
            todas_categorias = Categoria.objects.all().order_by('nome')
        elif has_polos_vinculados:
            todas_categorias = Categoria.objects.filter(
                Q(disponivel_todas=True) | Q(unidades__id__in=ids_permitidos)
            ).distinct().order_by('nome')
        else:
            todas_categorias = Categoria.objects.filter(disponivel_todas=True).order_by('nome')
    except Exception:
        todas_categorias = Categoria.objects.all().order_by('nome')

    categorias_out = todas_categorias.filter(tipo="OUT")
    categorias_in = todas_categorias.filter(tipo="IN")
    
    filtro_contas = Q(ativo=True)
    if is_gestor:
        filtro_contas &= (Q(permite_gestores=True) | Q(criado_por=user))
    elif is_controller:
        filtro_contas &= ((Q(permite_controller=True) & Q(somente_criador=False)) | Q(criado_por=user))
    else:
        filtro_contas &= ((Q(permite_operador=True) & Q(somente_criador=False)) | Q(criado_por=user))

    todas_contas_bancarias = ContaBancaria.objects.filter(filtro_contas).order_by("nome")

    context = {
        "perfil": perfil,
        "is_operador": is_operador,
        "is_controller": is_controller,
        "is_gestor": is_gestor,
        "pode_configurar": pode_configurar,
        "unidades_permitidas": list(unidades_permitidas),
        "todas_unidades_mapeadas": todas_unidades_mapeadas,
        "todas_unidades": todas_unidades,
        "todas_categorias": todas_categorias,
        "todas_contas_bancarias": todas_contas_bancarias,
        "categorias_out": categorias_out,
        "categorias_in": categorias_in,
        "unidades_selecionadas": [str(uid) for uid in unidades_get],
        "selected_categoria": int(categoria_id) if (categoria_id and categoria_id.isdigit()) else None,
        "selected_conta_origem": int(conta_origem_filtro) if (conta_origem_filtro and conta_origem_filtro.isdigit()) else None,
        "periodo": periodo,
        "data_de": data_de,
        "data_ate": data_ate,
        "hoje": hoje,
        "ano_atual": ano_atual,
        "mes_atual": mes_atual,
        "nome_mes_atual": nome_mes_atual,
        "mes_ant": mes_ant, "ano_ant": ano_ant,
        "mes_prox": mes_prox, "ano_prox": ano_prox,
        "despesas": despesas.order_by("data_vencimento"),
        "receitas": receitas.order_by("data_previsao"),
        "total_despesas": total_despesas,
        "total_custo_fixo": total_custo_fixo,
        "total_receitas": total_receitas,
        "saldo_projetado": saldo_projetado,
        "total_pendente": total_pendente,
        "total_aprovado": total_aprovado,
        "total_pago": total_pago,
    }
    return render(request, "financeiro/dashboard.html", context)

@login_required
def tendencia_view(request):
    user = request.user
    perfil = getattr(user, "perfil", None)
    is_gestor = user.is_superuser or user.is_staff or (perfil and perfil.papel == "GESTOR")
    is_controller = (perfil and perfil.papel == "CONTROLLER")

    if not (is_gestor or is_controller):
        messages.error(request, "Acesso restrito à tela de Tendência.")
        return redirect("dashboard_gestor")

    hoje = timezone.localdate()
    ano_param = request.GET.get("ano")
    mes_param = request.GET.get("mes")

    if ano_param and ano_param.isdigit() and mes_param and mes_param.isdigit():
        ano_atual = int(ano_param)
        mes_atual = int(mes_param)
    else:
        ano_atual = hoje.year
        mes_atual = hoje.month

    meses_nomes = {
        1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
    }
    nome_mes_atual = f"{meses_nomes.get(mes_atual, 'Mês')} {ano_atual}"

    if mes_atual == 1:
        mes_ant, ano_ant = 12, ano_atual - 1
    else:
        mes_ant, ano_ant = mes_atual - 1, ano_atual

    if mes_atual == 12:
        mes_prox, ano_prox = 1, ano_atual + 1
    else:
        mes_prox, ano_prox = mes_atual + 1, ano_atual

    # Dados dos últimos 6 meses para análise de tendência se período > 1 mês ou histórico recente
    tendencia_dados = []
    
    # Vamos gerar os últimos 6 meses até o mês atual selecionado para compor o gráfico de tendência
    curr_ano, curr_mes = ano_atual, mes_atual
    meses_analise = []
    for _ in range(6):
        meses_analise.insert(0, (curr_ano, curr_mes))
        if curr_mes == 1:
            curr_mes = 12
            curr_ano -= 1
        else:
            curr_mes -= 1

    for a, m in meses_analise:
        d_val = ContaPagar.objects.filter(data_vencimento__year=a, data_vencimento__month=m).aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")
        r_val = ContaReceber.objects.filter(data_previsao__year=a, data_previsao__month=m).aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")
        tendencia_dados.append({
            "mes_ano": f"{meses_nomes.get(m, '')[:3]}/{a}",
            "entradas": float(r_val),
            "saidas": float(d_val)
        })

    # Totais do mês atual selecionado para os cards inferiores (Foto 3)
    despesas_mes = ContaPagar.objects.filter(data_vencimento__year=ano_atual, data_vencimento__month=mes_atual)
    receitas_mes = ContaReceber.objects.filter(data_previsao__year=ano_atual, data_previsao__month=mes_atual)

    total_receitas = receitas_mes.aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")
    total_despesas = despesas_mes.aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")
    total_custo_fixo = despesas_mes.filter(recorrente=True).aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")
    saldo_projetado = total_receitas - total_despesas
    total_pendente = despesas_mes.filter(status="PENDENTE").aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")
    total_aprovado = despesas_mes.filter(status="APROVADO").aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")
    total_pago = despesas_mes.filter(status="PAGO").aggregate(Sum("valor"))["valor__sum"] or Decimal("0.00")

    context = {
        "is_gestor": is_gestor,
        "is_controller": is_controller,
        "ano_atual": ano_atual,
        "mes_atual": mes_atual,
        "nome_mes_atual": nome_mes_atual,
        "mes_ant": mes_ant, "ano_ant": ano_ant,
        "mes_prox": mes_prox, "ano_prox": ano_prox,
        "hoje": hoje,
        "tendencia_dados": json.dumps(tendencia_dados),
        "total_receitas": total_receitas,
        "total_despesas": total_despesas,
        "total_custo_fixo": total_custo_fixo,
        "saldo_projetado": saldo_projetado,
        "total_pendente": total_pendente,
        "total_aprovado": total_aprovado,
        "total_pago": total_pago,
    }
    return render(request, "financeiro/tendencia.html", context)

@login_required
def extrato_conta_view(request):
    user = request.user
    perfil = getattr(user, "perfil", None)
    is_gestor = user.is_superuser or user.is_staff or (perfil and perfil.papel == "GESTOR")
    is_controller = (perfil and perfil.papel == "CONTROLLER")

    if not (is_gestor or is_controller):
        messages.error(request, "Acesso restrito ao Extrato Bancário.")
        return redirect("dashboard_gestor")

    hoje = timezone.localdate()
    ano_param = request.GET.get("ano")
    mes_param = request.GET.get("mes")
    conta_id_param = request.GET.get("conta_id")

    if ano_param and ano_param.isdigit() and mes_param and mes_param.isdigit():
        ano_atual = int(ano_param)
        mes_atual = int(mes_param)
    else:
        ano_atual = hoje.year
        mes_atual = hoje.month

    meses_nomes = {
        1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
    }
    nome_mes_atual = f"{meses_nomes.get(mes_atual, 'Mês')} {ano_atual}"

    if mes_atual == 1:
        mes_ant, ano_ant = 12, ano_atual - 1
    else:
        mes_ant, ano_ant = mes_atual - 1, ano_atual

    if mes_atual == 12:
        mes_prox, ano_prox = 1, ano_atual + 1
    else:
        mes_prox, ano_prox = mes_atual + 1, ano_atual

    filtro_contas = Q(ativo=True)
    if is_gestor:
        filtro_contas &= (Q(permite_gestores=True) | Q(criado_por=user))
    else:
        filtro_contas &= ((Q(permite_controller=True) & Q(somente_criador=False)) | Q(criado_por=user))

    contas_disponiveis = ContaBancaria.objects.filter(filtro_contas).order_by("nome")

    conta_selecionada = None
    if conta_id_param and conta_id_param.isdigit():
        conta_selecionada = contas_disponiveis.filter(id=int(conta_id_param)).first()

    despesas_qs = ContaPagar.objects.select_related("unidade", "categoria", "conta_origem").filter(
        data_vencimento__year=ano_atual,
        data_vencimento__month=mes_atual
    )
    receitas_qs = ContaReceber.objects.select_related("unidade", "categoria", "conta_destino").filter(
        data_previsao__year=ano_atual,
        data_previsao__month=mes_atual
    )

    if conta_selecionada:
        despesas_qs = despesas_qs.filter(conta_origem=conta_selecionada)
        receitas_qs = receitas_qs.filter(conta_destino=conta_selecionada)

    itens_extrato = []

    for r in receitas_qs:
        itens_extrato.append({
            "tipo": "IN",
            "data": r.data_previsao,
            "unidade": r.unidade.nome,
            "descricao": r.origem + (f" - {r.descricao}" if r.descricao else ""),
            "categoria": r.categoria.nome if r.categoria else "-",
            "conta_nome": r.conta_destino.nome if r.conta_destino else "Não vinculada",
            "valor": r.valor,
            "status": r.get_status_display(),
            "realizado": (r.status == "CONFIRMADO"),
            "obj": r
        })

    for d in despesas_qs:
        itens_extrato.append({
            "tipo": "OUT",
            "data": d.data_vencimento,
            "unidade": d.unidade.nome,
            "descricao": d.fornecedor + (f" - {d.descricao}" if d.descricao else ""),
            "categoria": d.categoria.nome if d.categoria else "-",
            "conta_nome": d.conta_origem.nome if d.conta_origem else "Não vinculada",
            "valor": d.valor,
            "status": d.get_status_display(),
            "realizado": (d.status == "PAGO"),
            "obj": d
        })

    itens_extrato.sort(key=lambda x: x["data"])

    saldo_acumulado = Decimal("0.00")
    total_entradas = Decimal("0.00")
    total_saidas = Decimal("0.00")
    total_entradas_realizadas = Decimal("0.00")
    total_saidas_realizadas = Decimal("0.00")

    for item in itens_extrato:
        if item["tipo"] == "IN":
            total_entradas += item["valor"]
            saldo_acumulado += item["valor"]
            if item["realizado"]:
                total_entradas_realizadas += item["valor"]
        else:
            total_saidas += item["valor"]
            saldo_acumulado -= item["valor"]
            if item["realizado"]:
                total_saidas_realizadas += item["valor"]
        item["saldo_pos"] = saldo_acumulado

    saldo_final_projetado = total_entradas - total_saidas
    saldo_final_realizado = total_entradas_realizadas - total_saidas_realizadas

    context = {
        "is_gestor": is_gestor,
        "is_controller": is_controller,
        "ano_atual": ano_atual,
        "mes_atual": mes_atual,
        "nome_mes_atual": nome_mes_atual,
        "mes_ant": mes_ant, "ano_ant": ano_ant,
        "mes_prox": mes_prox, "ano_prox": ano_prox,
        "contas_disponiveis": contas_disponiveis,
        "conta_selecionada": conta_selecionada,
        "conta_id_param": conta_id_param,
        "itens_extrato": itens_extrato,
        "total_entradas": total_entradas,
        "total_saidas": total_saidas,
        "saldo_final_projetado": saldo_final_projetado,
        "total_entradas_realizadas": total_entradas_realizadas,
        "total_saidas_realizadas": total_saidas_realizadas,
        "saldo_final_realizado": saldo_final_realizado,
    }
    return render(request, "financeiro/extrato.html", context)

@login_required
def exportar_extrato_excel(request):
    ano = int(request.GET.get("ano", timezone.localdate().year))
    mes = int(request.GET.get("mes", timezone.localdate().month))
    conta_id = request.GET.get("conta_id")

    despesas_qs = ContaPagar.objects.select_related("unidade", "categoria", "conta_origem").filter(
        data_vencimento__year=ano,
        data_vencimento__month=mes
    )
    receitas_qs = ContaReceber.objects.select_related("unidade", "categoria", "conta_destino").filter(
        data_previsao__year=ano,
        data_previsao__month=mes
    )

    conta_nome = "Consolidado"
    if conta_id and conta_id.isdigit():
        cb = ContaBancaria.objects.filter(id=int(conta_id)).first()
        if cb:
            conta_nome = cb.nome
            despesas_qs = despesas_qs.filter(conta_origem=cb)
            receitas_qs = receitas_qs.filter(conta_destino=cb)

    itens = []
    for r in receitas_qs:
        itens.append({
            "tipo": "ENTRADA (Money IN)",
            "data": r.data_previsao,
            "unidade": r.unidade.nome,
            "descricao": r.origem + (f" - {r.descricao}" if r.descricao else ""),
            "categoria": r.categoria.nome if r.categoria else "-",
            "conta": r.conta_destino.nome if r.conta_destino else "-",
            "valor_in": float(r.valor),
            "valor_out": 0.0,
            "status": r.get_status_display()
        })

    for d in despesas_qs:
        itens.append({
            "tipo": "SAÍDA (Money OUT)",
            "data": d.data_vencimento,
            "unidade": d.unidade.nome,
            "descricao": d.fornecedor + (f" - {d.descricao}" if d.descricao else ""),
            "categoria": d.categoria.nome if d.categoria else "-",
            "conta": d.conta_origem.nome if d.conta_origem else "-",
            "valor_in": 0.0,
            "valor_out": float(d.valor),
            "status": d.get_status_display()
        })

    itens.sort(key=lambda x: x["data"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Extrato {mes:02d}-{ano}"

    headers = ["Data", "Tipo", "Unidade", "Descrição / Beneficiário", "Categoria", "Conta Bancária", "Entrada (R$)", "Saída (R$)", "Saldo Acumulado (R$)", "Status"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    saldo_acc = 0.0
    for it in itens:
        saldo_acc += (it["valor_in"] - it["valor_out"])
        ws.append([
            it["data"].strftime("%d/%m/%Y"),
            it["tipo"],
            it["unidade"],
            it["descricao"],
            it["categoria"],
            it["conta"],
            it["valor_in"] if it["valor_in"] > 0 else "",
            it["valor_out"] if it["valor_out"] > 0 else "",
            saldo_acc,
            it["status"]
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for col in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=extrato_{conta_nome}_{ano}_{mes:02d}.xlsx"
    wb.save(response)
    return response

@login_required
def gerenciar_contas_view(request):
    user = request.user
    perfil = getattr(user, "perfil", None)
    is_gestor = user.is_superuser or user.is_staff or (perfil and perfil.papel == "GESTOR")

    if not is_gestor:
        messages.error(request, "Acesso restrito ao Gestor Master.")
        return redirect("dashboard_gestor")

    if request.method == "POST":
        nome = request.POST.get("nome")
        banco = request.POST.get("banco", "")
        agencia = request.POST.get("agencia", "")
        conta = request.POST.get("conta", "")
        titular = request.POST.get("titular", "")
        chave_pix = request.POST.get("chave_pix", "")
        permite_gestores = request.POST.get("permite_gestores") == "on"
        permite_controller = request.POST.get("permite_controller") == "on"
        permite_operador = request.POST.get("permite_operador") == "on"
        somente_criador = request.POST.get("somente_criador") == "on"

        if nome:
            ContaBancaria.objects.create(
                nome=nome,
                banco=banco,
                agencia=agencia,
                conta=conta,
                titular=titular,
                chave_pix=chave_pix,
                permite_gestores=permite_gestores,
                permite_controller=permite_controller,
                permite_operador=permite_operador,
                somente_criador=somente_criador,
                criado_por=user
            )
            messages.success(request, f"Conta '{nome}' cadastrada com sucesso!")
        return redirect("gerenciar_contas")

    contas = ContaBancaria.objects.all().order_by("nome")
    return render(request, "financeiro/contas.html", {"contas": contas})

@login_required
def editar_conta_bancaria_view(request, pk):
    user = request.user
    perfil = getattr(user, "perfil", None)
    is_gestor = user.is_superuser or user.is_staff or (perfil and perfil.papel == "GESTOR")

    if not is_gestor:
        messages.error(request, "Acesso restrito ao Gestor Master.")
        return redirect("dashboard_gestor")

    conta_obj = get_object_or_404(ContaBancaria, pk=pk)

    if request.method == "POST":
        conta_obj.nome = request.POST.get("nome", conta_obj.nome)
        conta_obj.banco = request.POST.get("banco", "")
        conta_obj.agencia = request.POST.get("agencia", "")
        conta_obj.conta = request.POST.get("conta", "")
        conta_obj.titular = request.POST.get("titular", "")
        conta_obj.chave_pix = request.POST.get("chave_pix", "")
        conta_obj.permite_gestores = request.POST.get("permite_gestores") == "on"
        conta_obj.permite_controller = request.POST.get("permite_controller") == "on"
        conta_obj.permite_operador = request.POST.get("permite_operador") == "on"
        conta_obj.somente_criador = request.POST.get("somente_criador") == "on"
        conta_obj.save()
        messages.success(request, f"Conta '{conta_obj.nome}' atualizada com sucesso!")
        return redirect("gerenciar_contas")

    return redirect("gerenciar_contas")

@login_required
def deletar_conta_bancaria_view(request, pk):
    user = request.user
    perfil = getattr(user, "perfil", None)
    is_gestor = user.is_superuser or user.is_staff or (perfil and perfil.papel == "GESTOR")

    if not is_gestor:
        messages.error(request, "Acesso restrito ao Gestor Master.")
        return redirect("dashboard_gestor")

    if request.method == "POST":
        cb = get_object_or_404(ContaBancaria, pk=pk)
        cb.delete()
        messages.success(request, "Conta bancária excluída com sucesso.")
    return redirect("gerenciar_contas")

@login_required
def exportar_contas_mes(request):
    user = request.user
    perfil = getattr(user, "perfil", None)
    is_gestor = user.is_superuser or user.is_staff or (perfil and perfil.papel == "GESTOR")
    is_controller = (perfil and perfil.papel == "CONTROLLER")

    if not (is_gestor or is_controller):
        return redirect("dashboard_gestor")

    ano = int(request.GET.get("ano", timezone.localdate().year))
    mes = int(request.GET.get("mes", timezone.localdate().month))

    despesas = ContaPagar.objects.filter(data_vencimento__year=ano, data_vencimento__month=mes)
    receitas = ContaReceber.objects.filter(data_previsao__year=ano, data_previsao__month=mes)

    dados = {
        "versao": "uniges-4.7",
        "exportado_em": timezone.now().isoformat(),
        "ano": ano,
        "mes": mes,
        "despesas": [],
        "receitas": []
    }

    for d in despesas:
        dados["despesas"].append({
            "unidade_nome": d.unidade.nome,
            "categoria_nome": d.categoria.nome if d.categoria else None,
            "conta_origem_nome": d.conta_origem.nome if d.conta_origem else None,
            "fornecedor": d.fornecedor,
            "descricao": d.descricao,
            "dados_pix": d.dados_pix,
            "valor": str(d.valor),
            "data_vencimento": d.data_vencimento.isoformat(),
            "status": "PENDENTE",
            "recorrente": d.recorrente,
            "parcela_atual": d.parcela_atual,
            "total_parcelas": d.total_parcelas,
            "grupo_recorrencia_id": d.grupo_recorrencia_id
        })

    for r in receitas:
        dados["receitas"].append({
            "unidade_nome": r.unidade.nome,
            "categoria_nome": r.categoria.nome if r.categoria else None,
            "conta_destino_nome": r.conta_destino.nome if r.conta_destino else None,
            "origem": r.origem,
            "descricao": r.descricao,
            "valor": str(r.valor),
            "data_previsao": r.data_previsao.isoformat(),
            "status": "PREVISTO",
            "recorrente": r.recorrente,
            "parcela_atual": r.parcela_atual,
            "total_parcelas": r.total_parcelas,
            "grupo_recorrencia_id": r.grupo_recorrencia_id
        })

    response = HttpResponse(json.dumps(dados, ensure_ascii=False, indent=2), content_type="application/json")
    response["Content-Disposition"] = f"attachment; filename=uniges_contas_{ano}_{mes:02d}.json"
    return response

@login_required
def importar_contas_mes(request):
    user = request.user
    perfil = getattr(user, "perfil", None)
    is_gestor = user.is_superuser or user.is_staff or (perfil and perfil.papel == "GESTOR")
    is_controller = (perfil and perfil.papel == "CONTROLLER")

    if not (is_gestor or is_controller):
        return redirect("dashboard_gestor")

    if request.method == "POST":
        arquivo = request.FILES.get("arquivo_json")
        novo_ano = request.POST.get("novo_ano")
        novo_mes = request.POST.get("novo_mes")

        if not arquivo or not novo_ano or not novo_mes:
            messages.error(request, "Selecione o arquivo JSON e o mês/ano de destino.")
            return redirect("dashboard_gestor")

        try:
            conteudo = arquivo.read().decode("utf-8")
            dados = json.loads(conteudo)
            n_ano = int(novo_ano)
            n_mes = int(novo_mes)

            importados_desp = 0
            importados_rec = 0

            for item in dados.get("despesas", []):
                unidade, _ = Unidade.objects.get_or_create(nome=item["unidade_nome"])
                cat = None
                if item.get("categoria_nome"):
                    cat, _ = Categoria.objects.get_or_create(nome=item["categoria_nome"], defaults={"tipo": "OUT", "disponivel_todas": True})

                conta_origem = None
                if item.get("conta_origem_nome"):
                    conta_origem, _ = ContaBancaria.objects.get_or_create(nome=item["conta_origem_nome"])

                dt_antiga = datetime.fromisoformat(item["data_vencimento"]).date()
                dia = min(dt_antiga.day, 28)
                nova_data = dt_antiga.replace(year=n_ano, month=n_mes, day=dia)

                ContaPagar.objects.create(
                    unidade=unidade,
                    categoria=cat,
                    conta_origem=conta_origem,
                    fornecedor=item["fornecedor"],
                    descricao=item.get("descricao"),
                    dados_pix=item.get("dados_pix"),
                    valor=Decimal(item["valor"]),
                    data_vencimento=nova_data,
                    status="PENDENTE",
                    recorrente=item.get("recorrente", False),
                    parcela_atual=item.get("parcela_atual", 1),
                    total_parcelas=item.get("total_parcelas", 1),
                    grupo_recorrencia_id=item.get("grupo_recorrencia_id"),
                    criado_por=request.user
                )
                importados_desp += 1

            for item in dados.get("receitas", []):
                unidade, _ = Unidade.objects.get_or_create(nome=item["unidade_nome"])
                cat = None
                if item.get("categoria_nome"):
                    cat, _ = Categoria.objects.get_or_create(nome=item["categoria_nome"], defaults={"tipo": "IN", "disponivel_todas": True})

                conta_destino = None
                if item.get("conta_destino_nome"):
                    conta_destino, _ = ContaBancaria.objects.get_or_create(nome=item["conta_destino_nome"])

                dt_antiga = datetime.fromisoformat(item["data_previsao"]).date()
                dia = min(dt_antiga.day, 28)
                nova_data = dt_antiga.replace(year=n_ano, month=n_mes, day=dia)

                ContaReceber.objects.create(
                    unidade=unidade,
                    categoria=cat,
                    conta_destino=conta_destino,
                    origem=item["origem"],
                    descricao=item.get("descricao"),
                    valor=Decimal(item["valor"]),
                    data_previsao=nova_data,
                    status="PREVISTO",
                    recorrente=item.get("recorrente", False),
                    parcela_atual=item.get("parcela_atual", 1),
                    total_parcelas=item.get("total_parcelas", 1),
                    grupo_recorrencia_id=item.get("grupo_recorrencia_id")
                )
                importados_rec += 1

            messages.success(request, f"Importação realizada com sucesso! {importados_desp} despesas e {importados_rec} receitas importadas para {n_mes:02d}/{n_ano}.")
        except Exception as e:
            messages.error(request, f"Erro ao processar arquivo JSON: {e}")

    return redirect("dashboard_gestor")

@login_required
def gerenciar_categorias_view(request):
    user = request.user
    perfil = getattr(user, "perfil", None)
    is_gestor = user.is_superuser or user.is_staff or (perfil and perfil.papel == "GESTOR")
    is_controller = (perfil and perfil.papel == "CONTROLLER")

    if not (is_gestor or is_controller):
        return redirect("dashboard_gestor")

    nomes_uninter = ["Piracicaba", "Americana", "Sumaré", "Sumaré 2", "Rio Claro", "Araras", "Leme", "Conchal", "Pirassununga", "Matão", "Araraquara", "Olímpia", "Luciana Pessoal", "Rangélia Pessoal", "Alexandre", "Eletroposto Rosario"]

    for nome in nomes_uninter:
        Unidade.objects.get_or_create(nome=nome, defaults={"ativo": True})

    todas_unidades = list(Unidade.objects.filter(ativo=True))
    unidades_dict = {u.nome.lower(): u for u in todas_unidades}

    def map_unidades(nomes):
        return [unidades_dict.get(n.lower()) for n in nomes if n.lower() in unidades_dict]

    coluna_uninter = map_unidades(nomes_uninter)

    if request.method == "POST":
        nome = request.POST.get("nome")
        tipo = request.POST.get("tipo", "OUT")
        disponivel_todas = request.POST.get("disponivel_todas") == "on"
        unidades_selecionadas = request.POST.getlist("unidades_categoria")

        if nome:
            cat = Categoria.objects.create(
                nome=nome,
                tipo=tipo,
                disponivel_todas=disponivel_todas
            )
            if not disponivel_todas and unidades_selecionadas:
                cat.unidades.set(unidades_selecionadas)
            messages.success(request, f"Categoria '{nome}' criada com sucesso!")
        return redirect("gerenciar_categorias")

    categorias = Categoria.objects.prefetch_related("unidades").all().order_by("tipo", "nome")
    
    context = {
        "categorias": categorias,
        "coluna_uninter": coluna_uninter,
        "todas_unidades": todas_unidades,
    }
    return render(request, "financeiro/categorias.html", context)

@login_required
def editar_categoria_view(request, pk):
    user = request.user
    perfil = getattr(user, "perfil", None)
    is_gestor = user.is_superuser or user.is_staff or (perfil and perfil.papel == "GESTOR")
    is_controller = (perfil and perfil.papel == "CONTROLLER")

    if not (is_gestor or is_controller):
        return redirect("dashboard_gestor")

    cat = get_object_or_404(Categoria, pk=pk)

    if request.method == "POST":
        nome = request.POST.get("nome")
        tipo = request.POST.get("tipo", "OUT")
        disponivel_todas = request.POST.get("disponivel_todas") == "on"
        unidades_selecionadas = request.POST.getlist("unidades_categoria")

        if nome:
            cat.nome = nome
            cat.tipo = tipo
            cat.disponivel_todas = disponivel_todas
            cat.save()

            if disponivel_todas:
                cat.unidades.clear()
            else:
                cat.unidades.set(unidades_selecionadas)

            messages.success(request, f"Categoria '{cat.nome}' atualizada com sucesso!")
    return redirect("gerenciar_categorias")

@login_required
def deletar_categoria_view(request, pk):
    user = request.user
    perfil = getattr(user, "perfil", None)
    is_gestor = user.is_superuser or user.is_staff or (perfil and perfil.papel == "GESTOR")
    is_controller = (perfil and perfil.papel == "CONTROLLER")

    if (is_gestor or is_controller) and request.method == "POST":
        cat = get_object_or_404(Categoria, pk=pk)
        cat.delete()
        messages.success(request, f"Categoria '{cat.nome}' excluída.")
    return redirect("gerenciar_categorias")

@login_required
def criar_despesa_ajax(request):
    if request.method == "POST":
        unidade_id = request.POST.get("unidade")
        categoria_id = request.POST.get("categoria")
        conta_origem_id = request.POST.get("conta_origem")
        fornecedor = request.POST.get("fornecedor")
        descricao = request.POST.get("descricao")
        dados_pix = request.POST.get("dados_pix")
        valor = parse_currency(request.POST.get("valor", "0"))
        data_vencimento_str = request.POST.get("data_vencimento")
        boleto = request.FILES.get("boleto")
        aprovacao = request.FILES.get("aprovacao")
        comprovante = request.FILES.get("comprovante")
        
        is_recorrente = request.POST.get("is_recorrente") == "on"
        data_fim_str = request.POST.get("data_fim_recorrencia")

        status_inicial = "PAGO" if comprovante else "PENDENTE"
        dt_inicio = datetime.strptime(data_vencimento_str, "%Y-%m-%d").date()

        if is_recorrente:
            grupo_id = str(uuid.uuid4())[:12]
            if data_fim_str:
                dt_fim = datetime.strptime(data_fim_str, "%Y-%m-%d").date()
                datas_recorrencia = calcular_meses_datas(dt_inicio, dt_fim)
                total_parc = len(datas_recorrencia)
                for idx, dt in enumerate(datas_recorrencia, start=1):
                    ContaPagar.objects.create(
                        unidade_id=unidade_id,
                        categoria_id=categoria_id if categoria_id else None,
                        conta_origem_id=conta_origem_id if conta_origem_id else None,
                        fornecedor=fornecedor,
                        descricao=descricao,
                        dados_pix=dados_pix,
                        valor=valor,
                        data_vencimento=dt,
                        boleto=boleto if idx == 1 else None,
                        aprovacao=aprovacao if idx == 1 else None,
                        comprovante=comprovante if idx == 1 else None,
                        status=status_inicial if idx == 1 else "PENDENTE",
                        criado_por=request.user,
                        recorrente=True,
                        parcela_atual=idx,
                        total_parcelas=total_parc,
                        grupo_recorrencia_id=grupo_id
                    )
                messages.success(request, f"Despesa recorrente criada ({total_parc} parcelas)!")
            else:
                datas_recorrencia = gerar_proximas_datas(dt_inicio, qtd_meses=24)
                for idx, dt in enumerate(datas_recorrencia, start=1):
                    ContaPagar.objects.create(
                        unidade_id=unidade_id,
                        categoria_id=categoria_id if categoria_id else None,
                        conta_origem_id=conta_origem_id if conta_origem_id else None,
                        fornecedor=fornecedor,
                        descricao=descricao,
                        dados_pix=dados_pix,
                        valor=valor,
                        data_vencimento=dt,
                        boleto=boleto if idx == 1 else None,
                        aprovacao=aprovacao if idx == 1 else None,
                        comprovante=comprovante if idx == 1 else None,
                        status=status_inicial if idx == 1 else "PENDENTE",
                        criado_por=request.user,
                        recorrente=True,
                        parcela_atual=idx,
                        total_parcelas=0,
                        grupo_recorrencia_id=grupo_id
                    )
                messages.success(request, "Custo fixo mensal contínuo cadastrado!")
        else:
            ContaPagar.objects.create(
                unidade_id=unidade_id,
                categoria_id=categoria_id if categoria_id else None,
                conta_origem_id=conta_origem_id if conta_origem_id else None,
                fornecedor=fornecedor,
                descricao=descricao,
                dados_pix=dados_pix,
                valor=valor,
                data_vencimento=dt_inicio,
                boleto=boleto,
                aprovacao=aprovacao,
                comprovante=comprovante,
                status=status_inicial,
                criado_por=request.user,
                recorrente=False,
                parcela_atual=1,
                total_parcelas=1
            )
    return redirect(request.META.get("HTTP_REFERER", "dashboard_gestor"))

@login_required
def editar_despesa_ajax(request, pk):
    if request.method == "POST":
        despesa = get_object_or_404(ContaPagar, pk=pk)
        modo_edicao = request.POST.get("modo_edicao", "unico")

        unidade_id = request.POST.get("unidade")
        categoria_id = request.POST.get("categoria")
        conta_origem_id = request.POST.get("conta_origem")
        fornecedor = request.POST.get("fornecedor", despesa.fornecedor)
        descricao = request.POST.get("descricao", "")
        dados_pix = request.POST.get("dados_pix", "")
        valor = parse_currency(request.POST.get("valor", str(despesa.valor)))
        data_vencimento = request.POST.get("data_vencimento", despesa.data_vencimento)

        edit_recorrente = request.POST.get("edit_recorrente") == "on"
        tipo_rec = request.POST.get("edit_tipo_recorrencia", "parcelado")
        
        if edit_recorrente:
            if tipo_rec == "fixo":
                total_parc = 0
            else:
                t_parc = request.POST.get("edit_total_parcelas", "1")
                total_parc = int(t_parc) if t_parc.isdigit() else 1
        else:
            total_parc = 1

        if modo_edicao == "todos" and despesa.grupo_recorrencia_id:
            if total_parc > 0:
                ContaPagar.objects.filter(grupo_recorrencia_id=despesa.grupo_recorrencia_id, parcela_atual__gt=total_parc).delete()

            qs = ContaPagar.objects.filter(grupo_recorrencia_id=despesa.grupo_recorrencia_id).order_by('parcela_atual')
            for item in qs:
                if unidade_id:
                    item.unidade_id = unidade_id
                item.categoria_id = categoria_id if categoria_id else None
                item.conta_origem_id = conta_origem_id if conta_origem_id else None
                item.fornecedor = fornecedor
                item.descricao = descricao
                item.dados_pix = dados_pix
                item.valor = valor
                item.recorrente = edit_recorrente
                item.total_parcelas = total_parc
                if "boleto" in request.FILES:
                    item.boleto = request.FILES["boleto"]
                if "aprovacao" in request.FILES:
                    item.aprovacao = request.FILES["aprovacao"]
                if "comprovante" in request.FILES:
                    item.comprovante = request.FILES["comprovante"]
                    item.status = "PAGO"
                item.save()
            messages.success(request, "Toda a cadeia de parcelas foi atualizada com sucesso!")
        else:
            if unidade_id:
                despesa.unidade_id = unidade_id
            despesa.categoria_id = categoria_id if categoria_id else None
            despesa.conta_origem_id = conta_origem_id if conta_origem_id else None
            despesa.fornecedor = fornecedor
            despesa.descricao = descricao
            despesa.dados_pix = dados_pix
            despesa.valor = valor
            despesa.data_vencimento = data_vencimento
            despesa.recorrente = edit_recorrente
            despesa.total_parcelas = total_parc
            if edit_recorrente and tipo_rec != "fixo":
                p_atual = request.POST.get("edit_parcela_atual", "1")
                despesa.parcela_atual = int(p_atual) if p_atual.isdigit() else 1
            else:
                despesa.parcela_atual = 1

            if "boleto" in request.FILES:
                despesa.boleto = request.FILES["boleto"]
            if "aprovacao" in request.FILES:
                despesa.aprovacao = request.FILES["aprovacao"]
            if "comprovante" in request.FILES:
                despesa.comprovante = request.FILES["comprovante"]
                despesa.status = "PAGO"

            despesa.save()
            messages.success(request, "Despesa atualizada com sucesso!")
    return redirect(request.META.get("HTTP_REFERER", "dashboard_gestor"))

@login_required
def criar_receita_ajax(request):
    if request.method == "POST":
        user = request.user
        perfil = getattr(user, "perfil", None)
        is_gestor = user.is_superuser or user.is_staff or (perfil and perfil.papel == "GESTOR")
        is_controller = (perfil and perfil.papel == "CONTROLLER")
        
        if not (is_gestor or is_controller):
            return redirect("dashboard_gestor")

        unidade_id = request.POST.get("unidade")
        categoria_id = request.POST.get("categoria")
        conta_destino_id = request.POST.get("conta_destino")
        origem = request.POST.get("origem")
        descricao = request.POST.get("descricao")
        valor = parse_currency(request.POST.get("valor", "0"))
        data_previsao_str = request.POST.get("data_previsao")
        
        is_recorrente = request.POST.get("is_recorrente_rec") == "on"
        data_fim_str = request.POST.get("data_fim_recorrencia_rec")

        dt_inicio = datetime.strptime(data_previsao_str, "%Y-%m-%d").date()

        if is_recorrente:
            grupo_id = str(uuid.uuid4())[:12]
            if data_fim_str:
                dt_fim = datetime.strptime(data_fim_str, "%Y-%m-%d").date()
                datas_recorrencia = calcular_meses_datas(dt_inicio, dt_fim)
                total_parc = len(datas_recorrencia)
                for idx, dt in enumerate(datas_recorrencia, start=1):
                    ContaReceber.objects.create(
                        unidade_id=unidade_id,
                        categoria_id=categoria_id if categoria_id else None,
                        conta_destino_id=conta_destino_id if conta_destino_id else None,
                        origem=origem,
                        descricao=descricao,
                        valor=valor,
                        data_previsao=dt,
                        status="PREVISTO",
                        recorrente=True,
                        parcela_atual=idx,
                        total_parcelas=total_parc,
                        grupo_recorrencia_id=grupo_id
                    )
                messages.success(request, f"Receita recorrente criada ({total_parc} parcelas)!")
            else:
                datas_recorrencia = gerar_proximas_datas(dt_inicio, qtd_meses=24)
                for idx, dt in enumerate(datas_recorrencia, start=1):
                    ContaReceber.objects.create(
                        unidade_id=unidade_id,
                        categoria_id=categoria_id if categoria_id else None,
                        conta_destino_id=conta_destino_id if conta_destino_id else None,
                        origem=origem,
                        descricao=descricao,
                        valor=valor,
                        data_previsao=dt,
                        status="PREVISTO",
                        recorrente=True,
                        parcela_atual=idx,
                        total_parcelas=0,
                        grupo_recorrencia_id=grupo_id
                    )
                messages.success(request, "Receita fixa mensal contínua cadastrada!")
        else:
            ContaReceber.objects.create(
                unidade_id=unidade_id,
                categoria_id=categoria_id if categoria_id else None,
                conta_destino_id=conta_destino_id if conta_destino_id else None,
                origem=origem,
                descricao=descricao,
                valor=valor,
                data_previsao=dt_inicio,
                status="PREVISTO",
                recorrente=False,
                parcela_atual=1,
                total_parcelas=1
            )
    return redirect(request.META.get("HTTP_REFERER", "dashboard_gestor"))

@login_required
def editar_receita_ajax(request, pk):
    if request.method == "POST":
        user = request.user
        perfil = getattr(user, "perfil", None)
        is_gestor = user.is_superuser or user.is_staff or (perfil and perfil.papel == "GESTOR")
        is_controller = (perfil and perfil.papel == "CONTROLLER")
        
        if not (is_gestor or is_controller):
            return redirect("dashboard_gestor")

        receita = get_object_or_404(ContaReceber, pk=pk)
        modo_edicao = request.POST.get("modo_edicao", "unico")
        
        unidade_id = request.POST.get("unidade")
        categoria_id = request.POST.get("categoria")
        conta_destino_id = request.POST.get("conta_destino")
        origem = request.POST.get("origem", receita.origem)
        descricao = request.POST.get("descricao", "")
        valor = parse_currency(request.POST.get("valor", str(receita.valor)))
        data_previsao = request.POST.get("data_previsao", receita.data_previsao)
        
        edit_recorrente = request.POST.get("edit_recorrente_rec") == "on"
        tipo_rec = request.POST.get("edit_tipo_recorrencia_rec", "parcelado")
        
        if edit_recorrente:
            if tipo_rec == "fixo":
                total_parc = 0
            else:
                t_parc = request.POST.get("edit_total_parcelas_rec", "1")
                total_parc = int(t_parc) if t_parc.isdigit() else 1
        else:
            total_parc = 1

        if modo_edicao == "todos" and receita.grupo_recorrencia_id:
            if total_parc > 0:
                ContaReceber.objects.filter(grupo_recorrencia_id=receita.grupo_recorrencia_id, parcela_atual__gt=total_parc).delete()

            qs = ContaReceber.objects.filter(grupo_recorrencia_id=receita.grupo_recorrencia_id).order_by('parcela_atual')
            for item in qs:
                if unidade_id:
                    item.unidade_id = unidade_id
                item.categoria_id = categoria_id if categoria_id else None
                item.conta_destino_id = conta_destino_id if conta_destino_id else None
                item.origem = origem
                item.descricao = descricao
                item.valor = valor
                item.recorrente = edit_recorrente
                item.total_parcelas = total_parc
                item.save()
            messages.success(request, "Toda a cadeia de parcelas de receitas foi atualizada com sucesso!")
        else:
            if unidade_id:
                receita.unidade_id = unidade_id
            receita.categoria_id = categoria_id if categoria_id else None
            receita.conta_destino_id = conta_destino_id if conta_destino_id else None
            receita.origem = origem
            receita.descricao = descricao
            receita.valor = valor
            receita.data_previsao = data_previsao
            receita.recorrente = edit_recorrente
            receita.total_parcelas = total_parc
            if edit_recorrente and tipo_rec != "fixo":
                p_atual = request.POST.get("edit_parcela_atual_rec", "1")
                receita.parcela_atual = int(p_atual) if p_atual.isdigit() else 1
            else:
                receita.parcela_atual = 1

            receita.save()
            messages.success(request, "Receita atualizada com sucesso!")
    return redirect(request.META.get("HTTP_REFERER", "dashboard_gestor"))

@login_required
def deletar_despesa(request, pk):
    if request.method == "POST":
        despesa = get_object_or_404(ContaPagar, pk=pk)
        modo = request.POST.get("modo_exclusao", "unico")
        if modo == "todos" and despesa.grupo_recorrencia_id:
            total_removidos = ContaPagar.objects.filter(grupo_recorrencia_id=despesa.grupo_recorrencia_id).delete()[0]
            messages.success(request, f"Todas as {total_removidos} parcelas/ocorrências da despesa foram excluídas com sucesso.")
        else:
            despesa.delete()
            messages.success(request, "Despesa excluída com sucesso.")
    return redirect(request.META.get("HTTP_REFERER", "dashboard_gestor"))

@login_required
def deletar_receita(request, pk):
    if request.method == "POST":
        receita = get_object_or_404(ContaReceber, pk=pk)
        modo = request.POST.get("modo_exclusao", "unico")
        if modo == "todos" and receita.grupo_recorrencia_id:
            total_removidos = ContaReceber.objects.filter(grupo_recorrencia_id=receita.grupo_recorrencia_id).delete()[0]
            messages.success(request, f"Todas as {total_removidos} parcelas/ocorrências da receita foram excluídas com sucesso.")
        else:
            receita.delete()
            messages.success(request, "Receita excluída com sucesso.")
    return redirect(request.META.get("HTTP_REFERER", "dashboard_gestor"))

@login_required
def aprovar_em_lote(request):
    if request.method == "POST":
        user = request.user
        is_gestor = user.is_superuser or user.is_staff
        perfil = getattr(user, "perfil", None)
        is_controller = (perfil and perfil.papel == "CONTROLLER")

        acao = request.POST.get("acao")
        contas_ids = request.POST.getlist("contas")
        if contas_ids:
            if acao == "marcar_aprovado":
                ContaPagar.objects.filter(id__in=contas_ids).update(status="APROVADO")
            elif acao == "marcar_pago":
                ContaPagar.objects.filter(id__in=contas_ids).update(status="PAGO")
            elif acao == "marcar_pendente":
                ContaPagar.objects.filter(id__in=contas_ids).update(status="PENDENTE")
            elif acao == "deletar_lote" and (is_gestor or is_controller):
                ContaPagar.objects.filter(id__in=contas_ids).delete()
    return redirect(request.META.get("HTTP_REFERER", "dashboard_gestor"))

@login_required
def exportar_excel(request):
    unidades_get = request.GET.getlist("unidades")
    unidade_unica = request.GET.get("unidade")
    if unidade_unica and not unidades_get:
        unidades_get = [unidade_unica]

    ano = int(request.GET.get("ano", timezone.localdate().year))
    mes = int(request.GET.get("mes", timezone.localdate().month))

    despesas = ContaPagar.objects.select_related("unidade", "categoria", "conta_origem").filter(data_vencimento__year=ano, data_vencimento__month=mes)
    if unidades_get:
        ids_filtrar = [int(uid) for uid in unidades_get if str(uid).isdigit()]
        if ids_filtrar:
            despesas = despesas.filter(unidade_id__in=ids_filtrar)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Contas {mes:02d}-{ano}"

    headers = ["ID", "Unidade", "Fornecedor", "Descrição", "Recorrência", "Conta Origem", "Dados Pix", "Vencimento", "Valor (R$)", "Status", "Categoria"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for d in despesas:
        if d.recorrente:
            rec_str = "Custo Fixo" if d.total_parcelas == 0 else f"Parcela {d.parcela_atual}/{d.total_parcelas}"
        else:
            rec_str = "Único"

        ws.append([
            d.id,
            d.unidade.nome,
            d.fornecedor,
            d.descricao or "-",
            rec_str,
            d.conta_origem.nome if d.conta_origem else "-",
            d.dados_pix or "-",
            d.data_vencimento.strftime("%d/%m/%Y"),
            float(d.valor),
            d.get_status_display(),
            d.categoria.nome if d.categoria else "-"
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for col in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=uniges_contas_{ano}_{mes:02d}.xlsx"
    wb.save(response)
    return response
@login_required
def tendencia_view(request):
    return render(request, "financeiro/tendencia.html", {})
